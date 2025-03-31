"""Prices for the excel file"""

import os
import shutil
from openpyxl import load_workbook
from pathlib import Path
import win32com.client as win32

# Waardes voor sheet Prijspeil
MAANDELIJKSE_HUUR_B3 = 11000000  # Huurprijs per maand in euro
LEEGSTAND_B4 = 0.04  # Leegstand als percentage van de maandelijkse huur
JAARLIJKSE_EXPLOITATIEKOSTEN_B5 = 1300  # Exploitatiekosten per jaar in euro
GROOTONDERHOUD_PER_10_JAAR_B6 = 20000  # Groot onderhoudskosten per 10 jaar in euro
AANTAL_EXPLOITATIEJAREN_B7 = 20  # Aantal jaren waarop de exploitatiekosten worden berekend
HUIDIGE_VON_PRIJS_B8 = 200000  # Huidige von prijs in euro
HUURPRIJS_STIJGING_PER_JAAR_B9 = 0.015  # Jaarlijkse huurprijsstijging (1.5%)
KOSTEN_STIJGING_PER_JAAR_B10 = 0.02  # Jaarlijkse stijging van overige kosten (2%)
WAARDEONTWIKKELING_VON_PER_JAAR_B11 = 0.5  # Jaarlijkse waardeontwikkeling van de von prijs (50%)
KOSTEN_KOPER_B12 = 0.04  # Kosten koper als percentage van de aankoopprijs
IRR_B13 = 0.06  # Internal Rate of Return (interne rentevoet)

# Bepaal het pad naar het Excel-bestand
BASE_DIR = Path(__file__).resolve().parent
EXCEL_PATH = BASE_DIR.parent / "data" / "sample_excel.xlsx"  # Pad naar het originele Excel-bestand

# Controleren of het Excel-bestand bestaat
if not EXCEL_PATH.exists():
    raise FileNotFoundError(
        f"Het bestand {EXCEL_PATH} bestaat niet. Zorg ervoor dat het bestand aanwezig is."
    )
else:
    print(f"Het bestand {EXCEL_PATH} bestaat.")


def update_excel_file():
    # Maak een kopie van het originele bestand
    COPY_PATH = BASE_DIR.parent / "data" / "kopie.xlsx"
    # Zorg dat de doelmap bestaat
    COPY_PATH.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(EXCEL_PATH, COPY_PATH)
    print(f"Kopie gemaakt van {EXCEL_PATH} naar {COPY_PATH}")

    # Open de gekopieerde Excel (met formules)
    wb = load_workbook(COPY_PATH)

    # Controleer en selecteer de juiste sheets
    if "Prijspeil" not in wb.sheetnames:
        raise ValueError("De sheet 'Prijspeil' bestaat niet in het Excel-bestand.")
    sheet_prijspeil = wb["Prijspeil"]
    if "Cashflow" not in wb.sheetnames:
        raise ValueError("De sheet 'Cashflow' bestaat niet in het Excel-bestand.")
    sheet_dcf = wb["Cashflow"]

    # Bewerk de sheet: schrijf aangepaste data naar cel A1
    sheet_prijspeil["A1"] = "Aangepaste data in een andere sheet"
    print("Cel A1 is aangepast.")

    # Update de cellen met de corresponderende variabelen
    sheet_prijspeil["B3"] = MAANDELIJKSE_HUUR_B3
    sheet_prijspeil["B4"] = LEEGSTAND_B4
    sheet_prijspeil["B5"] = JAARLIJKSE_EXPLOITATIEKOSTEN_B5
    sheet_prijspeil["B6"] = GROOTONDERHOUD_PER_10_JAAR_B6
    sheet_prijspeil["B7"] = AANTAL_EXPLOITATIEJAREN_B7
    sheet_prijspeil["B8"] = HUIDIGE_VON_PRIJS_B8
    sheet_prijspeil["B9"] = HUURPRIJS_STIJGING_PER_JAAR_B9
    sheet_prijspeil["B10"] = KOSTEN_STIJGING_PER_JAAR_B10
    sheet_prijspeil["B11"] = WAARDEONTWIKKELING_VON_PER_JAAR_B11
    sheet_prijspeil["B12"] = KOSTEN_KOPER_B12
    sheet_prijspeil["B13"] = IRR_B13
    print("De cellen met de variabele waarden zijn aangepast.")

    # Sla de wijzigingen op in het kopiebestand
    wb.save(COPY_PATH)
    print(f"Wijzigingen opgeslagen in {COPY_PATH}")

    # Exporteer de sheet "Cashflow" als PDF via de COM-interface
    pdf_path = COPY_PATH.with_suffix(".pdf")
    export_sheet_to_pdf(COPY_PATH, pdf_path, sheet_name="Cashflow")
    print(f"PDF geëxporteerd naar {pdf_path}")


def export_sheet_to_pdf(excel_file, pdf_file, sheet_name="Cashflow"):
    """
    Exporteert de opgegeven sheet uit een Excel-bestand als PDF.
    Vereist: Windows en een geïnstalleerde versie van Excel.
    """
    excel_app = win32.gencache.EnsureDispatch("Excel.Application")
    excel_app.Visible = False
    try:
        wb = excel_app.Workbooks.Open(str(excel_file))
        # Forceer een recalculatie van de formules
        wb.RefreshAll()
        excel_app.CalculateUntilAsyncQueriesDone()
        # Selecteer de gewenste sheet
        ws = wb.Worksheets(sheet_name)
        # Exporteer de sheet als PDF (0 staat voor PDF)
        ws.ExportAsFixedFormat(0, str(pdf_file))
    finally:
        wb.Close(False)
        excel_app.Quit()


def main():
    print("Start van het script.")
    update_excel_file()


if __name__ == "__main__":
    main()
    print("Het script is uitgevoerd.")
